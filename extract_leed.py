#!/usr/bin/env python3
"""
Extract LEED v4 energy performance data from the Bechtel/Grimes
LEED submittal calculator spreadsheet into a small JSON the front-end
can consume.

Source:
    `1. Minimum_Optimize Energy Performance/02-Clarifications/
     20240529_Bechtel_LEEDv4MinEnergyPerfCalc.xlsm`
    Sheet: Performance_Outputs_Summary

Output:
    data/bechtel_leed.json

Adds the "design intent" layer to the digital twin:
- Per-end-use ASHRAE 90.1-2010 baseline (code minimum)
- Per-end-use proposed (as-designed, eQuest energy model output)
- Building metadata (sq ft, LEED ID, certification path)
- BMO meter → LEED end-use mapping for live performance gap analysis

Re-run whenever an updated LEED submittal arrives.

Usage:
    pip install openpyxl --break-system-packages
    python extract_leed.py path/to/calculator.xlsm
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl: pip install openpyxl")

STOP_TERMS = {'Electricity', 'Natural Gas', 'District Cooling',
              'District Heating', 'Energy model subtotal'}

# Meter mapping — informed by UCB Bechtel electrical design (M76 = 480/277V mechanical,
# M77 = 208/120V plug+lighting, M3 = main service tie).
METER_MAP = {
    '_doc': 'Maps BMO meter IDs to LEED end-use categories. M76 = 480/277V mechanical, '
            'M77 = 208/120V plug+lighting, M3 = main service.',
    '76': ['Space heating', 'Space cooling', 'Pumps', 'Heat rejection',
           'Fans - interior ventilation', 'Fans Exhaust'],
    '77': ['Interior lighting', 'Exterior lighting', 'Receptacle equipment',
           'Service water heating', 'IT equipment', 'Elevators and escalators'],
    '3': '_all_',
}


def num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def extract(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=False)

    # End-uses from Performance_Outputs_Summary
    ws = wb['Performance_Outputs_Summary']
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(rows)
                   if r and any(c == 'End_Use' for c in r if c))

    end_uses = []
    current = None
    for r in rows[hdr_idx + 1:hdr_idx + 60]:
        if not r or not any(r):
            continue
        end_use = r[1] if len(r) > 1 else None
        energy_type = r[4] if len(r) > 4 else None
        units_str = r[5] if len(r) > 5 else None
        baseline = r[6] if len(r) > 6 else None
        proposed = r[7] if len(r) > 7 else None
        savings = r[8] if len(r) > 8 else None

        if end_use and any(t in str(end_use) for t in STOP_TERMS):
            break
        if end_use and 'Total' in str(end_use):
            break

        if end_use and str(end_use).strip() and num(baseline) is not None:
            current = {
                'end_use': str(end_use).strip(),
                'energy_type': str(energy_type).strip() if energy_type else None,
                'baseline_kwh': round(num(baseline) or 0, 1),
                'proposed_kwh': round(num(proposed) or 0, 1),
                'savings_pct': round((num(savings) or 0) * 100, 1),
            }
            end_uses.append(current)
        elif (current and not end_use and num(baseline) is not None
              and units_str and 'Demand' in str(units_str)):
            current['baseline_kw'] = round(num(baseline) or 0, 2)
            current['proposed_kw'] = round(num(proposed) or 0, 2)

    # Drop zero-zero rows (end-uses not applicable to this project)
    end_uses = [eu for eu in end_uses
                if (eu['baseline_kwh'] or eu['proposed_kwh'])]

    # Building metadata from General Information
    gi = wb['General Information']
    info = {}
    for r in gi.iter_rows(min_row=1, max_row=60, values_only=True):
        if not r:
            continue
        label = None
        for c in r:
            if c is None:
                continue
            s = str(c).strip()
            if not label and s and not isinstance(c, (int, float)):
                label = s
            elif label and c is not None:
                info[label[:80]] = c
                break

    # Totals
    total_b = sum(eu['baseline_kwh'] for eu in end_uses)
    total_p = sum(eu['proposed_kwh'] for eu in end_uses)
    total_kw_b = sum(eu.get('baseline_kw', 0) for eu in end_uses)
    total_kw_p = sum(eu.get('proposed_kw', 0) for eu in end_uses)
    sqft = info.get('Conditioned building area (sq ft)', 72108)

    return {
        '_doc': 'LEED v4 EAp2/EAc1 energy performance data for UCB Bechtel '
                'Engineering Center (Grimes Hall). Extracted from LEED submittal '
                'spreadsheet (Performance_Outputs_Summary sheet). baseline = '
                'ASHRAE 90.1-2010 code minimum; proposed = as-designed energy '
                'model output (eQuest).',
        '_source_file': Path(xlsm_path).name,
        'building': {
            'leed_project_id': str(info.get('LEED Project ID #', '')),
            'name': info.get('LEED Project Name', ''),
            'rating_system': info.get('Rating system', ''),
            'conditioned_area_sqft': sqft,
            'unconditioned_area_sqft': info.get('Unconditioned building area (sq ft)'),
            'percent_new_construction': info.get('Percent new construction (%)*'),
            'percent_renovation_existing': info.get('Percent renovation/existing (%)*'),
        },
        'totals': {
            'baseline_annual_kwh': round(total_b, 0),
            'proposed_annual_kwh': round(total_p, 0),
            'baseline_peak_kw': round(total_kw_b, 1),
            'proposed_peak_kw': round(total_kw_p, 1),
            'design_savings_pct': round((total_b - total_p) / total_b * 100, 1) if total_b else 0,
            'eui_proposed_kwh_per_sqft_yr': round(total_p / sqft, 2) if sqft else None,
            'eui_proposed_kbtu_per_sqft_yr': round(total_p / sqft * 3.412, 1) if sqft else None,
        },
        'end_uses': end_uses,
        'meter_end_use_map': METER_MAP,
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('xlsm', help='Path to LEEDv4MinEnergyPerfCalc.xlsm')
    ap.add_argument('--output', default='data/bechtel_leed.json')
    args = ap.parse_args()

    data = extract(args.xlsm)
    with open(args.output, 'w') as f:
        json.dump(data, f, indent=2)

    t = data['totals']
    print(f"  {len(data['end_uses'])} end-uses extracted")
    print(f"  Baseline: {t['baseline_annual_kwh']:,.0f} kWh/yr")
    print(f"  Proposed: {t['proposed_annual_kwh']:,.0f} kWh/yr "
          f"({t['design_savings_pct']:+.1f}% savings)")
    print(f"  EUI proposed: {t['eui_proposed_kwh_per_sqft_yr']} kWh/sqft/yr "
          f"(≈ {t['eui_proposed_kbtu_per_sqft_yr']} kBtu/sqft/yr)")
    print(f"→ wrote {args.output}")


if __name__ == '__main__':
    main()
